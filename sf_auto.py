#!/usr/bin/env Python
# coding=utf-8




import os
import time
import random
import openpyxl
import csv
import xlrd
import pyautogui
from selenium import webdriver
from openpyxl import Workbook
from tkinter import *

username = '珠海大手印科技'
userpassw ='zhgsy2017'
options = webdriver.ChromeOptions()
# 本处为下载文件储存位置，后继需要改为自动选择路径。
prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': os.getcwd()}
options.add_experimental_option('prefs', prefs)

st4_name = 'a'
jdurl = 'https://passport.jd.com/new/login.aspx?rs=vc&ReturnUrl=//vcp.jd.com'
jd_page = webdriver.Chrome(chrome_options=options)
jd_page.get(jdurl)
jd_page.maximize_window()
jd_page.implicitly_wait(20)
nowpage = 0
roottk = Tk()
def main_page():
    pass


def jd_loginpage():
    global nowpage
    if nowpage == 0:
        nowpage = 1;
    jd_action()


# # 人工登录京东主页
def log_jd():
    jd_page.find_element_by_link_text("账户登录").click()
    jd_page.find_element_by_id('loginname').clear()
    jd_page.find_element_by_id('loginname').send_keys(username)
    time.sleep(1)
    jd_page.find_element_by_id('nloginpwd').clear()
    jd_page.find_element_by_id('nloginpwd').send_keys(userpassw)
    roottk.title('登录后点击我')
    roottk.wm_attributes('-topmost',1)
    Button(roottk,text = '请登录jd后点击我', width = 30, height = 5, command = jd_loginpage).pack()
    roottk.mainloop()
    # jd_page.find_element_by_id('loginsubmit').click()
    # jd_loginpage()


def jd_action():
    roottk.destroy()
    # while nowpage == 1:
    time.sleep(2)
    try:
        if(jd_page.find_element_by_class_name('ant-modal-mask')):
            # jd_page.find_element_by_link_text("知道了").click()
            jd_page.find_element_by_xpath('/html/body/div/div/div[2]/div/div[1]/div[3]/button').click()
    except:
        print("第一层遮罩错误，请联系开发维护")

    time.sleep(2)
    jd_page.find_element_by_link_text('厂家直送').click()

    time.sleep(5)
    jd_page.find_element_by_link_text("订单生产与跟踪").click()

    time.sleep(4)
    jd_page.find_element_by_id("outboundTab").click()

    # 点击左下角出库
    time.sleep(2)
    jd_page.find_element_by_id("cb_outboundGrid").click()

    time.sleep(2)
    jd_page.find_element_by_id("outbound_doExport").click()

    # 点击确定
    time.sleep(5)
    inner_page = jd_page.find_element_by_id('system_confirm')
    inner_page.find_element_by_xpath('//*[@id="system_confirm"]/div[3]/button[1]').click()
    inner_table = jd_page.find_element_by_id('outboundGrid')
    inner_tablenum = inner_table.find_elements_by_tag_name('tr')
    if(int(len(inner_tablenum))>2):
        time.sleep(2)
        jd_page.find_element_by_class_name("ui-pg-div").click()

        time.sleep(5)
        inner_page = jd_page.find_element_by_id('system_confirm')
        inner_page.find_element_by_xpath('//*[@id="system_confirm"]/div[3]/button[1]').click()
    else:
        print('当前无可选项')

    time.sleep(3)
    jd_page.find_element_by_id("deliveryTab").click()

    time.sleep(3)
    jd_page.find_element_by_id("delivery_doExport").click()

    # 点击确认导出发货表
    time.sleep(5)
    jd_page.find_element_by_xpath('//*[@id="system_confirm"]/div[3]/button[1]/span[2]').click()

    jd_page.minimize_window()
    # excel_auto()
    jd_step2()

def jd_step2():

    list_name = os.listdir()
    global open_fileN
    for i in list_name:
        if(re.match(r'02(\s*)(\S*).xlsx$',i)):
            open_fileN = re.match(r'02(\s*)(\S*).xlsx$',i).group()

    wb1 = Workbook()
    st1 = wb1.active
    st4_namelst = ('04'+str(time.strftime('%y-%m-%d'))+r'-1sffhdc-lsyd.xls')
    wb2 = xlrd.open_workbook(st4_namelst)
    st2 = wb2.sheet_by_index(0)
    from_array = ['订单号','配送公司','运单号']
    for k in range(3):
        st1.cell(row = 1, column = k+1).value = from_array[k]
    # for i in range(st2.max_row):
    for j in range(5):
        for k in range(3):
            from_name = from_array[k]
            if(from_name == "配送公司"):
                for i in range(st2.nrows-1):
                    input_value = "顺丰快递"
                    st1.cell(row = i+2,column = k+1).value = input_value
            if(st2.cell_value(0,j) == from_name):
                for i in range(st2.nrows-1):
                    input_value = st2.cell_value(i+1, j)
                    st1.cell(row = i+2,column = k+1).value = input_value

    wb05_name = 'a05-ddaoru.xls'
    wb1.save(wb05_name)
    wb1_pash = os.path.abspath(wb05_name)
    pyautogui.typewrite(wb1_pash,interval= 0.1)
    pyautogui.press('enter')
    jd_page.maximize_window()
    #deliveryGrid_page_left > table > tbody > tr > td:nth-child(4) > div
    for i in range(10):
        time.sleep(random.random())
        pyautogui.scroll(-60*i)
    table_div = jd_page.find_element_by_id('pg_deliveryGrid_page')
    tablem = table_div.find_element_by_tag_name('tbody')
    tbtrs = tablem.find_elements_by_tag_name('td')
    for i in range(len(tbtrs)):
        if(tbtrs[i].text.strip() == '批量发货'):
            tbtrs[i].click()
            break

    # tablem.find_element_by_link_text("批量发货").click()
    time.sleep(2)
    jd_page.find_element_by_id('excelFile').click()
    pyautogui.press('capslock')
    pyautogui.typewrite(wb1_pash,interval=0.1)
    pyautogui.press('enter')

    time.sleep(2)
    # jd_page.find_element_by_id('batchSubmit').click()

    wblst = openpyxl.load_workbook(open_fileN)
    stlst = wblst['data']
    total_col = stlst.max_column
    plus_col = ["优化备注",'快递运单','发货人','备注']
    for i in range(len(plus_col)):
        stlst.cell(row = 1, column = total_col -1+i).value = plus_col[i]
    for x in range(st1.max_row):
        first_xv = st1.cell(row = x+2, column =1).value
        for y in range(st1.max_row):
            mmv=stlst.cell(row = y+2, column = 1).value
            if ( first_xv != None and mmv != None and int(first_xv) - int(mmv) == 0):
                tt1 = st1.cell(row = x+2,column =3).value
                tt2 = stlst.cell(row = y+2, column = total_col).value
                stlst.cell(row = y+2, column = total_col).value = tt1
    wb1.close()
    wb06_name = '06 汇总表'+str(time.strftime('%y-%m-%d'))+r'.xls'
    wblst.save(wb06_name)
    wblst.close()
    # here will add some words
    jd_page.find_element_by_id('batchSubmit').click()

# excel操作函数
def excel_auto():
    # 第一步，将csv文件转换为xlsx文件
    list_name = os.listdir()
    for i in list_name:
        if(re.match(r'[\u53d1](\w*).csv$',i)):
            open_filename = re.match(r'[\u53d1](\w*).csv$',i).group()
    time.sleep(3)
    fname = open(open_filename,'r')

    # csv = pandas.read_csv(fname, encoding='utf-8')
    lines = csv.reader(fname)
    b2_xlsname = ('02 jd-dc-'+str(time.strftime('%y-%m-%d')+r'.xlsx'))

    workbook = Workbook()
    wst1 = workbook.active
    wst1.title = 'data'

    xxy = 1
    for line in lines:
        x = 1
        for i in line:
            wst1.cell(row = xxy, column = x).value = i
            x += 1
        xxy += 1
    fname.close()
    workbook.save(b2_xlsname)
    st1_copy_array =["客户订单号","收货人姓名","收货人电话","收货人地址","商品数量"]
    st2_past_array=["用户订单号","联系人","联系电话","收件详细地址","托寄物数量"]
    st2_guding = {"收件公司":".","付款方式":"寄付月结","托寄物品":"数码产品","托寄物内容":"耗材","件数":"1","业务类型":"顺丰特惠"}
    # 第二步,将表2对应数据导入表3
    wb1 = openpyxl.load_workbook(b2_xlsname)
    wb2 = openpyxl.load_workbook(r'03CommonTemplate2_vip.xlsx')
    st1 = wb1['data']
    st2 = wb2['导入数据表']

    st1_rownums = st1.max_row
    st1_colnums = st1.max_column

    st2_rownums = st2.max_row
    st2_colnums = st2.max_column


    while len(st1_copy_array)>0:
        for y in range(st1_colnums):
            first_value = st1.cell(row = 1, column = y+1).value
            if(first_value ==st1_copy_array[0]):
                st1_copy_array.remove(first_value)
                break
        while len(st2_past_array)> 0:
            for ii in range(st2_colnums):
                two_value2 = st2.cell(row = 2, column = ii+1).value
                if(two_value2 in st2_guding):
                    for i2 in range(st1_rownums):
                        st2.cell(row = i2+3, column = ii+1).value = st2_guding.get(two_value2)
            for i in range(st2_colnums):
                two_value = st2.cell(row = 2, column = i+1).value
                if(two_value == None):
                    two_value = st2.cell(row = 1, column = i+1).value
                if(two_value == st2_past_array[0]):
                    st2_past_array.remove(two_value)
                    break
            break
    # 取对应的表2数值写入表3对应位置
        for rn in range(st1_rownums):
            target_value = st1.cell(row = rn+1, column = y+1).value
            st2.cell(row = rn+3, column = i+1).value = target_value
        else:
            continue


    st1_columns_name = st1.iter_rows();
    wb2.save('test2.xlsx')
    wb1.close()
    wb2.close()

    sf_auto()

    # excel_auto()
    # 顺丰界面操作函数
def sf_auto():
    roottk2 = Tk()
    roottk2.wm_attributes('-topmost',1)
    roottk2.title("请登录顺丰大客户发件系统")
    roottk2.geometry('800x300+800+300')
    global st4_name
    def sf_autorun():
        roottk2.destroy()
        # 设置程序延迟时间，本处为随机值。
        pyautogui.PAUSE =2+random.random()
        # 保护措施，当程序出现错误，将鼠标拉倒左上角，则可以退出。
        pyautogui.FAILSAFE = True

        max_size = pyautogui.size()
        print(max_size)
        # 点击订单新建
        # time.sleep(2)
        ddxj_where = pyautogui.locateCenterOnScreen(r'lc-sf\ddxj.png')
        if ddxj_where == None:
            # ddxj_where =(61,119)
            ddxj_where =(61,115)#1440-900
            # ddxj_where =(90,150)
            # ddxj_where = (max_size[0]*4.68765*0.01,max_size[1]*13.8888*0.01)
        pyautogui.moveTo(ddxj_where)
        pyautogui.click()
        # 点击标准模板导入
        # time.sleep(1)
        bzmbdr_where = pyautogui.locateCenterOnScreen(r'lc-sf\bzmbdr.png')
        if bzmbdr_where == None:
            # bzmbdr_where =(183,109)
            bzmbdr_where =(185,110)#1440-900
            # bzmbdr_where =(270,135)
            # bzmbdr_where =(max_size[0]*14.0625*0.01,max_size[1]*12.5*0.01)
        pyautogui.moveTo(bzmbdr_where)
        pyautogui.click()
        time.sleep(3)
        # 输入待导入文件路径
        st3_path = os.path.abspath('test2.xlsx')
        pyautogui.typewrite(st3_path, interval=0.1)
        time.sleep(2)
        # 点击确定，导入文件
        pyautogui.press("enter")
        time.sleep(3)
        # 提交，不打印
        tj_where = pyautogui.locateCenterOnScreen(r'lc-sf\tj.png')
        if tj_where == None:
            # tj_where =(1000,990)
            # tj_where =(1020,950)
            tj_where =(775,810)#1140-900
            # tj_where =(max_size[0]*52.0833*0.01,max_size[1]*87.963*0.01)
        pyautogui.moveTo(tj_where)
        pyautogui.click()
        # 确认是否需要打印
        while True:
            if(pyautogui.locateOnScreen(r'lc-sf\sfdy.png') or pyautogui.locateCenterOnScreen(r'lc-sf\sfdy2.png') or pyautogui.locateCenterOnScreen(r'lc-sf\sfdy3.png')):
                # pyautogui.click(1120,540)
                pyautogui.press('right')
                # time.sleep(0.5)
                pyautogui.press('space')
                # pyautogui.click()
                break
        # 订单管理
        time.sleep(3)
        ddgl_where = pyautogui.locateCenterOnScreen(r'lc-sf\ddgl.png')
        if ddgl_where == None:
            # ddgl_where =(63,860)
            # ddgl_where =(90,780)
            ddgl_where =(60,675)#1440-900
            # ddgl_where =(max_size[0]*4.6875*0.01,max_size[1]*72.2223*0.01)
        pyautogui.moveTo(ddgl_where)
        pyautogui.click()
        # 订单查询打印
        time.sleep(1)
        ddcxdy_where = pyautogui.locateCenterOnScreen(r'lc-sf\ddcxdy.png')
        if ddcxdy_where == None:
            # ddcxdy_where =(60,145)
            # ddcxdy_where =(90,190)
            ddcxdy_where =(60,140)
            # ddcxdy_where =(max_size[0]*4.6875*0.01,max_size[1]*17.5925*0.01)
        pyautogui.moveTo(ddcxdy_where)
        pyautogui.click()

        time.sleep(1)
        # dyzt_where = (740,110)
        # dyzt_where = (770,145)
        dyzt_where = (700,110)
        # dyzt_where = (max_size[0]*40.1041*0.01,max_size[1]*13.4259*0.01)
        pyautogui.moveTo(dyzt_where)
        pyautogui.click()
        time.sleep(1)
        # pyautogui.moveRel(0,60)
        pyautogui.press('down')
        pyautogui.press('enter')
        # pyautogui.click()
        time.sleep(1)
        # pyautogui.moveTo(180,240)
        # pyautogui.moveTo(245,300)
        pyautogui.moveTo(180,240)
        # pyautogui.moveTo(max_size[0]*12.7604*0.01,max_size[1]*27.7765*0.01)
        pyautogui.click()
        time.sleep(1)
        pyautogui.moveRel(170,0)
        pyautogui.click()
        global st4_name
        time.sleep(1)
        st4_path = (os.getcwd()+r'\04'+str(time.strftime('%y-%m-%d'))+r'-1sffhdc-lsyd.xls')
        st4_name = st4_path
        pyautogui.typewrite(st4_path,interval = 0.1)
        pyautogui.press('enter')
        time.sleep(1)
        # pyautogui.click(1120,540)
        pyautogui.press('enter')

        time.sleep(2)
        jd_step2()

    Button(roottk2,text = "登录成功顺丰大客户发件系统后点击我",height = 5,command = sf_autorun, bg = 'red').pack()
    roottk2.mainloop()

def main():
    log_jd()#主入口



main()
